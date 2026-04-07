"""
02_fetch_research_explorer.py

Phase 1, step 2: pull the broader NASA Space Station Research Explorer
(SSRE) catalog. SSRE is the canonical NASA catalogue of every investigation
ever conducted aboard ISS — across all five partner agencies and every
discipline (biology, human research, physical science, technology,
education) — and is roughly 3x the size of OSDR's omics-focused subset.

INVESTIGATION (Spec 01, Step 1)
-------------------------------
Tried four access methods before settling on the download approach:

  1. Direct JSON XHR endpoint behind https://www.nasa.gov/mission/station/
     research-explorer/ — page is a static HTML wrapper, no XHR/JSON
     endpoint, no client-side fetch. Dead end.
  2. NASA TechPort API — exposes funded NASA "projects", not individual
     ISS investigations. Wrong granularity.
  3. **Downloadable XLSX reports** — the SSRE landing page links three
     official, refreshed-daily Excel files:
        - All_Experiments_Report.xlsx   (~2,925 investigations)
        - All_Publications_Report.xlsx  (~8,200 publications)
        - All_Facilities_Report.xlsx
     They are clean, tabular, and require no API key. **This is what we
     use.** It is the simplest, most reliable, and most maintainable
     option — and it's the same data source the SSRE site itself reads.
  4. HTML scraping / SerpAPI — unnecessary given option 3 works.

The XLSX schema only carries: Short Name, Full Name, Principal
Investigators, Developers, Expeditions, Category, Sponsoring Agency.
There are no objectives/approach/results fields available, so those are
left empty per the spec ("do not make up data"). To give the downstream
classifier something to work with, every SSRE row is enriched with the
titles of any publications linked to that investigation in the
publications report — this materially improves keyword classification
on bare-title experiments.

Output:
  - data/raw/research_explorer.json     (raw rows from both XLSX files)
  - data/processed/research_explorer.csv (flattened, OSDR-schema columns)

After fetching, this script also runs the merge step (Spec 01 Step 3):
  - Backs up data/processed/osdr_experiments.csv → osdr_experiments_backup_pre_ssre.csv
  - Deduplicates SSRE against OSDR by (title + principal_investigator)
  - Writes the combined catalog back to data/processed/osdr_experiments.csv

Re-running this script is safe: the XLSX is re-downloaded each run, and
the merge always restores from backup if one already exists.

See SPACE_HEALTH_SPECS.md section 2.2 / 3.2 and specs/01_add_ssre_experiments.md.
"""

from __future__ import annotations

import csv
import re
import sys
import time
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

sys.path.insert(0, str(Path(__file__).resolve().parent))
from config import (  # noqa: E402
    CHECKPOINT_DIR,
    PROCESSED_DIR,
    RAW_DIR,
    REQUEST_HEADERS,
    save_json,
)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
SSRE_BASE = "https://www.nasa.gov/mission_pages/station/research/experiments/explorer"
EXPERIMENTS_XLSX_URL = f"{SSRE_BASE}/All_Experiments_Report.xlsx"
PUBLICATIONS_XLSX_URL = f"{SSRE_BASE}/All_Publications_Report.xlsx"

CHECKPOINT_FILE = CHECKPOINT_DIR / "research_explorer_checkpoint.json"
RAW_OUTPUT = RAW_DIR / "research_explorer.json"
RAW_EXPERIMENTS_XLSX = RAW_DIR / "ssre_All_Experiments_Report.xlsx"
RAW_PUBLICATIONS_XLSX = RAW_DIR / "ssre_All_Publications_Report.xlsx"

CSV_OUTPUT = PROCESSED_DIR / "research_explorer.csv"
OSDR_CSV = PROCESSED_DIR / "osdr_experiments.csv"
OSDR_BACKUP = PROCESSED_DIR / "osdr_experiments_backup_pre_ssre.csv"

# OSDR column schema — must match scripts/01_fetch_nasa_osdr.py exactly
CSV_COLUMNS = [
    "osID",
    "title",
    "objectives",
    "approach",
    "results",
    "sponsoringAgency",
    "researchAreas",
    "nasaPrograms",
    "factors",
    "publications_count",
    "publication_titles",
    "principal_investigator",
    "pi_institution",
    "all_people",
    "releaseDate",
    "source_url",
]

LOG_EVERY = 25
SSRE_PAGE_URL = "https://www.nasa.gov/mission/station/research-explorer/"


# ---------------------------------------------------------------------------
# HTTP session
# ---------------------------------------------------------------------------
def make_session() -> requests.Session:
    session = requests.Session()
    retry = Retry(
        total=4,
        backoff_factor=1.5,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset(["GET"]),
    )
    session.mount("https://", HTTPAdapter(max_retries=retry))
    session.headers.update(REQUEST_HEADERS)
    return session


def download_xlsx(session: requests.Session, url: str, dest: Path) -> Path:
    print(f"[ssre] GET {url}")
    try:
        resp = session.get(url, timeout=120)
        resp.raise_for_status()
    except Exception as exc:
        raise RuntimeError(f"Failed to download {url}: {exc}") from exc
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(resp.content)
    print(f"[ssre]   saved {len(resp.content):,} bytes → {dest.name}")
    return dest


# ---------------------------------------------------------------------------
# XLSX parsing
# ---------------------------------------------------------------------------
def _clean(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    # Collapse the trailing semicolon-newline pattern that NASA uses
    # everywhere ("Foo, Bar;\nBaz, Qux;\n" → "Foo, Bar; Baz, Qux")
    s = s.replace("\r", "").replace("\n", " ")
    s = re.sub(r"\s+;", ";", s)
    s = re.sub(r"\s+", " ", s)
    s = s.strip().rstrip(";").strip()
    return s


def _split_semicolon_list(value: str) -> list[str]:
    if not value:
        return []
    return [p.strip() for p in value.split(";") if p.strip()]


def _first_person_name(pi_field: str) -> str:
    """
    SSRE PI fields look like:
        "Jane Doe, Ph.D., NASA Ames Research Center, Moffett Field; Other Person, ..."
    Extract just the first person's name (everything before the first comma+title).
    Falls back to the full first segment if no recognizable pattern.
    """
    if not pi_field:
        return ""
    first_segment = pi_field.split(";")[0].strip()
    # Drop trailing institution: split on comma and keep the leading
    # name parts. Heuristic: take chunks until we hit one that looks
    # like a degree (contains "Ph.D", "M.D", "M.S", "Dr.") or an
    # institution keyword.
    parts = [p.strip() for p in first_segment.split(",")]
    name_parts: list[str] = []
    for p in parts:
        low = p.lower()
        if any(t in low for t in ("ph.d", "m.d", "m.s.", "dr.", "msc", "phd")):
            break
        # Institution keywords
        if any(
            t in low
            for t in (
                "university",
                "institute",
                "center",
                "centre",
                "agency",
                "nasa",
                "esa",
                "jaxa",
                "csa",
                "company",
                "inc",
                "corporation",
                "laboratory",
                "lab.",
                "gmbh",
                "ltd",
                "llc",
            )
        ):
            break
        name_parts.append(p)
        if len(name_parts) >= 3:  # safety: never grab more than 3 chunks
            break
    return ", ".join(name_parts).strip()


def _first_institution(developer_field: str) -> str:
    """First entry of the developer list is a reasonable proxy for PI institution."""
    if not developer_field:
        return ""
    return developer_field.split(";")[0].strip()


def parse_experiments_xlsx(path: Path) -> list[dict[str, Any]]:
    """
    Parse the All_Experiments_Report.xlsx file. The first 3 rows are
    metadata (title, generated date, blank); row 3 (0-indexed) is the
    header; data starts at row 4.
    """
    print(f"[ssre] Parsing {path.name}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 5:
        raise RuntimeError(f"{path} has too few rows ({len(rows)}) — format changed?")

    header = [str(h or "").strip() for h in rows[3]]
    expected_header = [
        "Short Name",
        "Full Name",
        "Principal Investigator(s)",
        "Developer(s)",
        "Expedition(s)",
        "Category",
        "Sponsoring Space Agency",
    ]
    if header[: len(expected_header)] != expected_header:
        print(f"[ssre]   WARNING: header mismatch. Got: {header}")

    raw_rows: list[dict[str, Any]] = []
    for raw in rows[4:]:
        if not raw or not raw[0]:
            continue
        short_name = _clean(raw[0])
        if not short_name:
            continue
        raw_rows.append(
            {
                "short_name": short_name,
                "full_name": _clean(raw[1]) or short_name,
                "principal_investigators": _clean(raw[2]),
                "developers": _clean(raw[3]),
                "expeditions": _clean(raw[4]),
                "category": _clean(raw[5]),
                "sponsoring_agency": _clean(raw[6]),
            }
        )
    print(f"[ssre]   parsed {len(raw_rows)} investigations")
    return raw_rows


def parse_publications_xlsx(path: Path) -> dict[str, list[str]]:
    """
    Group citations by Short Name. Returns {short_name: [citation, ...]}.
    """
    print(f"[ssre] Parsing {path.name}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    by_exp: dict[str, list[str]] = defaultdict(list)
    for raw in rows[4:]:
        if not raw or not raw[0]:
            continue
        short_name = _clean(raw[0])
        citation = _clean(raw[2])
        if short_name and citation:
            by_exp[short_name].append(citation)
    print(f"[ssre]   {len(by_exp)} investigations have linked publications")
    return dict(by_exp)


# ---------------------------------------------------------------------------
# Normalization to OSDR schema
# ---------------------------------------------------------------------------
_SLUG_RE = re.compile(r"[^a-zA-Z0-9]+")


def _slug(short_name: str) -> str:
    s = _SLUG_RE.sub("-", short_name).strip("-").upper()
    return s or "UNKNOWN"


def normalize_to_osdr_schema(
    raw_rows: list[dict[str, Any]],
    pubs_by_exp: dict[str, list[str]],
) -> list[dict[str, Any]]:
    """
    Map SSRE rows onto the OSDR column layout. Empty objectives/approach/
    results — SSRE doesn't carry those fields. We use:

      researchAreas  ← SSRE Category
      nasaPrograms   ← "Expedition <X>" string
      pi_institution ← first developer entry
      all_people     ← list of PI names joined by "; "
      source_url     ← SSRE landing page (no per-investigation URL exists)

    SSRE has a small number of duplicate Short Names (e.g. "BIOKIN 4"
    and "BIOKIN-4"). We disambiguate them by appending "-2", "-3", ... so
    that every emitted osID is unique downstream.
    """
    out: list[dict[str, Any]] = []
    seen_ids: dict[str, int] = {}
    for raw in raw_rows:
        short = raw["short_name"]
        slug = _slug(short)
        base_id = f"SSRE-{slug}"
        if base_id in seen_ids:
            seen_ids[base_id] += 1
            os_id = f"{base_id}-{seen_ids[base_id]}"
        else:
            seen_ids[base_id] = 1
            os_id = base_id

        pi_field = raw["principal_investigators"]
        pi_name = _first_person_name(pi_field)
        all_pis = _split_semicolon_list(pi_field)
        all_pi_names = [_first_person_name(p) for p in all_pis if p]
        all_pi_names = [n for n in all_pi_names if n]

        pubs = pubs_by_exp.get(short, [])
        pub_titles = " | ".join(pubs)

        expeditions = raw["expeditions"]
        nasa_programs = f"ISS Expedition {expeditions}" if expeditions else "ISS"

        out.append(
            {
                "osID": os_id,
                "title": raw["full_name"],
                "objectives": "",
                "approach": "",
                "results": "",
                "sponsoringAgency": raw["sponsoring_agency"],
                "researchAreas": raw["category"],
                "nasaPrograms": nasa_programs,
                "factors": "",
                "publications_count": len(pubs),
                "publication_titles": pub_titles,
                "principal_investigator": pi_name,
                "pi_institution": _first_institution(raw["developers"]),
                "all_people": "; ".join(all_pi_names),
                "releaseDate": "",
                "source_url": SSRE_PAGE_URL,
            }
        )
    return out


# ---------------------------------------------------------------------------
# CSV I/O
# ---------------------------------------------------------------------------
def write_csv(rows: list[dict[str, Any]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=CSV_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        return [dict(r) for r in reader]


# ---------------------------------------------------------------------------
# Merge / dedup against OSDR
# ---------------------------------------------------------------------------
def _dedup_key(row: dict[str, Any]) -> str:
    title = (row.get("title") or "").strip().lower()
    pi = (row.get("principal_investigator") or "").strip().lower()
    # Strip degrees and trailing institution from PI for matching
    pi = re.sub(r"\s*,\s*(ph\.?d\.?|m\.?d\.?|m\.?s\.?|dr\.?|msc|phd)\b.*", "", pi)
    pi = pi.strip().rstrip(",").strip()
    title = re.sub(r"\s+", " ", title)
    return f"{title}|{pi}"


def merge_into_osdr(
    ssre_rows: list[dict[str, Any]],
) -> tuple[int, int, int]:
    """
    Backup → load OSDR → build dedup key set → drop SSRE rows whose
    (title+PI) collide with OSDR → write the combined CSV back to
    osdr_experiments.csv.

    Returns (osdr_count, ssre_added, dropped_duplicates).
    """
    if not OSDR_CSV.exists():
        raise RuntimeError(f"{OSDR_CSV} missing — run scripts/01_fetch_nasa_osdr.py first")

    if OSDR_BACKUP.exists():
        print(f"[ssre] Backup already exists at {OSDR_BACKUP.name} — leaving it untouched")
    else:
        OSDR_BACKUP.write_bytes(OSDR_CSV.read_bytes())
        print(f"[ssre] Backed up {OSDR_CSV.name} → {OSDR_BACKUP.name}")

    # Always merge from the BACKUP so re-runs are idempotent — otherwise
    # we'd accumulate SSRE rows on top of an already-merged file.
    osdr_rows = read_csv(OSDR_BACKUP)
    print(f"[ssre] Loaded {len(osdr_rows)} OSDR rows from backup")

    osdr_keys = {_dedup_key(r) for r in osdr_rows}

    kept: list[dict[str, Any]] = []
    dropped = 0
    for r in ssre_rows:
        key = _dedup_key(r)
        if key in osdr_keys:
            dropped += 1
            continue
        kept.append(r)

    combined = osdr_rows + kept
    write_csv(combined, OSDR_CSV)
    print(
        f"[ssre] Merge complete: {len(osdr_rows)} OSDR + {len(kept)} new SSRE "
        f"({dropped} dupes dropped) = {len(combined)} total"
    )
    return len(osdr_rows), len(kept), dropped


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    session = make_session()

    # Step 1: download both XLSX files (raw artifacts go to data/raw/)
    try:
        download_xlsx(session, EXPERIMENTS_XLSX_URL, RAW_EXPERIMENTS_XLSX)
        time.sleep(0.5)
        download_xlsx(session, PUBLICATIONS_XLSX_URL, RAW_PUBLICATIONS_XLSX)
    except Exception as exc:
        print(f"[ssre] FATAL: download failed: {exc}")
        raise SystemExit(1) from exc

    # Step 2: parse
    raw_rows = parse_experiments_xlsx(RAW_EXPERIMENTS_XLSX)
    pubs_by_exp = parse_publications_xlsx(RAW_PUBLICATIONS_XLSX)

    # Save raw JSON snapshot (for audit / future re-parsing without the XLSX)
    save_json(
        RAW_OUTPUT,
        {
            "source": EXPERIMENTS_XLSX_URL,
            "fetched_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
            "experiments": raw_rows,
            "publications_index": {
                k: v for k, v in pubs_by_exp.items()
            },
        },
    )
    print(f"[ssre]   raw  → {RAW_OUTPUT}")

    # Step 3: normalize and log progress every LOG_EVERY rows
    normalized = normalize_to_osdr_schema(raw_rows, pubs_by_exp)
    for idx, row in enumerate(normalized, start=1):
        if idx % LOG_EVERY == 0 or idx == len(normalized):
            print(f"[ssre] Fetched {idx}/{len(normalized)}: {row['osID']}")

    write_csv(normalized, CSV_OUTPUT)
    print(f"[ssre]   csv  → {CSV_OUTPUT}")

    # Save resume checkpoint (XLSX is single-shot, so the "checkpoint" is
    # just a marker that the fetch step completed)
    save_json(
        CHECKPOINT_FILE,
        {
            "fetched_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
            "experiment_count": len(normalized),
            "publications_indexed": len(pubs_by_exp),
            "method": "xlsx_download",
            "experiments_url": EXPERIMENTS_XLSX_URL,
            "publications_url": PUBLICATIONS_XLSX_URL,
        },
    )

    # Step 4: merge into OSDR catalog
    osdr_count, ssre_added, dropped = merge_into_osdr(normalized)

    print()
    print("=" * 60)
    print("[ssre] DONE")
    print("=" * 60)
    print(f"  SSRE investigations fetched : {len(normalized)}")
    print(f"  SSRE with publications      : {len(pubs_by_exp)}")
    print(f"  OSDR rows preserved         : {osdr_count}")
    print(f"  SSRE rows added (post-dedup): {ssre_added}")
    print(f"  Duplicates dropped          : {dropped}")
    print(f"  Combined catalog total      : {osdr_count + ssre_added}")
    print(f"  Combined CSV                : {OSDR_CSV}")
    print()
    print("Next: run `python scripts/05_classify_experiments.py` to classify the new rows.")


if __name__ == "__main__":
    main()
